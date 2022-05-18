import sched, time
from win32com.client import gencache
# from .regsvr import get_clsid
# from regsvr import get_clsid
# from regsvr import get_serv_list
import pathlib
from log.LOGS import LOGS
import os
import json
import xlrd
import xlwt
import pickle
import openpyxl
import win32com
import time
import dcom_da.regsvr

value_types = ['str', 'int', 'bool', 'float']


class DA_CLIENT:
    def __init__(self, host='127.0.0.1', server_name='Matrikon.OPC.Simulation.1', mode='SERVER', file=None, sheet=None,
                 MonitorHandler=None, UpdateRate=500):
        self._server = None
        self._browser = None
        self._opcGroupM = None
        self._opcGroupBuff = None
        self._monitorEventHandler = MonitorHandler
        self.monitorItemsID = []
        self._serverHandles = []
        self.restart_connection = False
        self._isPingSuccess = False
        self.log_restart_connection = 0
        self.host = host
        self.server_name = server_name
        self.isConnected = False
        self.inputFile = file
        self.inputFileSheet = sheet
        self.UpdateRate = UpdateRate
        self.Tree = None
        self.Mode = mode
        self.s = sched.scheduler(time.time, time.sleep)

    def GetTree(self):
        if self.Tree is not None:
            return self.Tree

        try:
            if self.isConnected == False:
                return None

            if self._server is not None:
                self._browser = self._server.CreateBrowser()

                if self._browser is not None:
                    if self.Mode == 'SERVER':
                        LOGS('dcom_da/DA_CLIENT.GetTree', 'Trying to read tags in mode "SERVER"', 'INFO')
                        if not os.path.exists('DA_TREE'):
                            LOGS('dcom_da/DA_CLIENT.GetTree', 'Create DA_TREE.json and writting tree', 'INFO')
                            self._browser.AccessRights = 0
                            self._opcGroupBuff = self._server.OPCGroups.Add('BufferGroup')
                            self.Tree = self._GetTreeItemByBranches()

                            with open('DA_TREE.json', 'w', encoding='utf-8') as f:
                                json.dump(obj=self.Tree, fp=f, indent=4, ensure_ascii=False, default=str)
                        else:
                            LOGS('dcom_da/DA_CLIENT.GetTree', 'Load tree from DA_TREE.json', 'INFO')
                            with open('DA_TREE.json', 'r') as file:
                                self.Tree = json.load(file)
                    else:
                        LOGS('dcom_da/DA_CLIENT.GetTree', 'Trying to read tags in mode "FILE"', 'INFO')
                        self.GetItemsFromFile()


        except Exception as err:
            LOGS('dcom_da/DA_CLIENT.GetTree', 'ERROR: In getting the tag tree', 'ERROR')
            print('GetTree Error::', err)

        return self.Tree

    def _GetTreeItemByBranches(self, branches=None, name=None):
        result = []

        if branches is None:
            self._browser.MoveToRoot()
            self._browser.ShowBranches()
        else:
            self._browser.MoveTo(branches)
            self._browser.ShowBranches()

        brancheCount = self._browser.Count
        for i in range(brancheCount):
            i += 1
            if branches is None:
                self._browser.MoveToRoot()
                self._browser.ShowBranches()
                branches2 = [self._browser.Item(i)]
            else:
                self._browser.MoveTo(branches)
                self._browser.ShowBranches()
                branches2 = [None] * (len(branches) + 1)

                for j in range(len(branches)):
                    branches2[j] = branches[j]
                branches2[-1] = self._browser.Item(i)

            if name == None:
                name_buff = ''
            else:
                name_buff = name + self._browser.Item(i) + '.'

            new_branch = {'Name': self._browser.Item(i),
                          'Type': 'folder',
                          'BrancheArray': self._GetTreeItemByBranches(branches2, name_buff),
                          'LeafArray': []}
            self._browser.MoveTo(branches2)
            self._browser.ShowLeafs()
            for j in range(self._browser.Count):
                j += 1
                new_leaf = {'Name': name_buff + self._browser.Item(j),
                            'Type': 'value'}
                try:
                    property_id = [1, 2]
                    property_id.insert(0, 0)

                    values, errors = self._server.GetItemProperties(name_buff + self._browser.Item(j),
                                                                    len(property_id) - 1,
                                                                    property_id)
                    if values[1].__class__.__name__ not in value_types:
                        continue
                    new_leaf['Value'] = values[1]
                except:
                    continue

                self.monitorItemsID.append(name_buff + self._browser.Item(j))
                new_branch['LeafArray'].append(new_leaf)
            result.append(new_branch)
        return result

    def FormMonitorItemList(self):
        tree = self.GetTree()

        def getLeafArray(tree):
            for branch in tree:
                if branch['Type'] == 'folder':
                    for leaf in branch['LeafArray']:
                        self.monitorItemsID.append(leaf['Name'])
                    getLeafArray(branch['BrancheArray'])
                else:
                    self.monitorItemsID.append(branch['Name'])

        getLeafArray(tree)

        # print(self.monitorItemsID)

    def SaveMonitorItemList(self):
        new_excel = xlwt.Workbook()
        ws = new_excel.add_sheet('output_sheet')
        ws.write(0, 0, 'instrumenttag')

        i = 1
        for tag in self.monitorItemsID:
            ws.write(i, 0, tag)
            i += 1

        new_excel.save('MonitorItemTags.xls')

    def AddItemId(self):
        try:
            if self._opcGroupM is not None:
                if self._opcGroupM.OPCItems.Count == 0:
                    self._serverHandles = [None] * len(self.monitorItemsID)
                    for i in range(len(self.monitorItemsID)):
                        try:
                            self._serverHandles[i] = self._opcGroupM.OPCItems.AddItem(self.monitorItemsID[i],
                                                                                      i).ServerHandle
                        except:
                            pass

                return True
        except Exception as err:
            LOGS('dcom_da/DA_CLIENT.AddItemId', 'Error: Adding new ItemId', 'ERROR')
            print('AddItemId Error::', err)

        return False

    def GetItemsFromFile(self):
        def read_sheet(sheet, file_fotmat='xls'):
            if file_fotmat == 'xlsx':
                ncols = sheet.max_column
                nrow = sheet.max_row
                colidx = dict((sheet.cell(row=1, column=i).value, i) for i in range(1, ncols + 1))
                tags = [sheet.cell(row=i, column=colidx["instrumenttag"]).value for i in range(1, nrow + 1)]
                tags = tags[1:]
            else:
                ncols = sheet.ncols
                nrow = sheet.nrows
                colidx = dict((sheet.cell(0, i).value, i) for i in range(ncols))
                tags = [sheet.cell(i, colidx["instrumenttag"]).value for i in range(1, nrow)]

            size = len(tags)

            for tag in tags:

                new_leaf = {'Name': tag,
                            'Type': 'value'}
                try:
                    property_id = [1, 2]
                    property_id.insert(0, 0)

                    values, errors = self._server.GetItemProperties(tag,
                                                                    len(property_id) - 1,
                                                                    property_id)
                    if values[1].__class__.__name__ not in value_types:
                        continue

                    new_leaf['Value'] = values[1]
                    new_leaf['Type_value'] = type(values[1]).__name__
                    # self.monitorItemsID.append(tag)
                except Exception as err:
                    print('Add item error:: cannot add a tag: [ {} ]  '.format(tag), err)
                    continue
                self.Tree.append(new_leaf)

        self.Tree = []

        if pathlib.Path(self.inputFile).suffix == '.xls':
            wb = xlrd.open_workbook(self.inputFile)
            sheet = wb.sheet_by_name(self.inputFileSheet)
            read_sheet(sheet, 'xls')

        elif pathlib.Path(self.inputFile).suffix == '.xlsx':

            wb = openpyxl.load_workbook(self.inputFile)
            sheet = wb.active
            read_sheet(sheet, 'xlsx')

    def StartMonitor(self, handlerInit):
        try:
            self.isMonitor = True
            if self.isConnected == False:
                print('connected false')
                return

            if self._server is not None:
                if self._opcGroupM is not None:
                    handler = win32com.client.WithEvents(self._opcGroupM, self._monitorEventHandler)
                    handlerInit(handler)
                    if self.AddItemId():
                        self._opcGroupM.UpdateRate = self.UpdateRate
                        self._opcGroupM.IsActive = True
                        self._opcGroupM.IsSubscribed = True
                    else:
                        print('add item is false')
                else:
                    print('group is none')
            else:
                print('server is None')
        except Exception as err:
            print('Start monitor Error::', err)

    def PingHost(self):
        response = os.system("ping " + self.host)
        if response == 0:
            return True
        else:
            return False

    def CheckConnected(self) -> object:
        result = False
        try:
            ping = self.PingHost()
            if ping == True:
                self._isPingSuccess = True
                if self._server != None:
                    serverState = int(self._server.ServerState)
                    if serverState == int(win32com.client.constants.OPCRunnig):
                        result = True
            else:
                self._isPingSuccess = False
                self._server = None
        except Exception as err:
            print('CheckConnected error::', err)
            self._isPingSuccess = False
            self._server = None
            a = [err]
            if err in a:
                result = False
        return result

    def restart_connection(self):
        # self.s.enter(5, 1, self.restart_connection)
        print(self.CheckConnected())

    def Connect(self):
        self.s.enter(20, 1, self.Connect)

        if not self.isConnected:
            attempt = 0
            while True:
                try:
                    self.CheckConnected()
                    if attempt == 1:
                        LOGS('dcom_da/DA_CLIENT.Connect',
                             'WARNING: Check if it works DA server DA_HOST: {}'.format(self.host), 'WARNING')
                    LOGS('dcom_da/DA_CLIENT.Connect', 'Trying connect to DA server... DA_HOST:{}'.format(self.host),
                         'INFO')
                    if (self._isPingSuccess and self.isConnected == False):
                        dll = win32com.client.gencache.EnsureModule(dcom_da.regsvr.get_clsid(), 0, 1, 0)
                        self._server = dll.OPCServer()
                        self._server.Connect(self.server_name, self.host)
                        LOGS('dcom_da/DA_CLIENT.Connect',
                             'Success: Connect to DA server! DA_HOST: {}'.format(self.host), 'INFO')
                        print('Successfully connected to DA Server on the host: {}'.format(self.host))

                        self._opcGroupM = win32com.client.Dispatch(self._server.OPCGroups.Add('MonitorGroup'))

                        self.isConnected = True

                    break
                except Exception as err:
                    self._server = None
                    LOGS('dcom_da/DA_CLIENT.Connect', 'Error: Can not connect к DA_HOST: {}'.format(self.host),
                         'ERROR')
                    print('Failed connect to DA server. Check the settings (server name, host or access rights)\n',
                          err)
                    attempt += 1
                time.sleep(5)
        else:
            try:
                self._server.Connect(self.server_name, self.host)
                if (self.restart_connection == True) and (self.log_restart_connection < 1):
                    self.log_restart_connection += 1
                    LOGS('dcom_da/DA_CLIENT.Connect', 'Server is working', 'INFO')
            except Exception as err:
                LOGS('dcom_da/DA_CLIENT.Connect', 'Error: Disconnected from DA SERVER, DA_HOST: {}'.format(self.host),
                     'ERROR')
                self.restart_connection = True
                print('Server is down')

    def Disconnect(self):
        try:
            # self.CheckConnected()
            # if self._isPingSuccess and self.isConnected:
            #     if self._server is not None:
            #         self._server.OPCGroups.RemoveAll()
            #         self._server.Disconnect()
            #         print('Successfully disconnected')

            self._server = None
            self._browser = None
            self._opcGroupM = None
            self.isConnected = False
        except Exception as err:
            print('Failed disconnect:', err)
